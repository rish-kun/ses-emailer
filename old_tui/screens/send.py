"""
Send Screen for SES Email TUI application.
Handles the actual sending of emails with progress tracking.
"""

import asyncio
import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from textual.app import ComposeResult
from textual.containers import Container, Horizontal, Vertical
from textual.screen import Screen
from textual.widgets import (
    Button,
    Footer,
    Header,
    Log,
    ProgressBar,
    Static,
)

from config.settings import format_source_email, get_email_address


class SendScreen(Screen):
    """Email sending screen with progress tracking."""

    BINDINGS = [
        ("escape", "cancel", "Cancel"),
    ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.is_sending: bool = False
        self.is_paused: bool = False
        self.cancel_requested: bool = False
        self.total_sent: int = 0
        self.total_failed: int = 0
        self.failed_emails: list[dict] = []  # List of {email, error} dicts
        self.current_batch_index: int = 0  # Track position for resume
        self.sending_complete: bool = False
        self.current_email_id: str | None = None  # Track email ID for database

    def compose(self) -> ComposeResult:
        yield Header()
        with Container(id="send-container"):
            yield Static(
                "[bold]>[/] Send Emails",
                classes="screen-title",
            )

            with Vertical(id="send-summary"):
                yield Static(
                    "Email Summary",
                    classes="section-header",
                )
                yield Static("", id="summary-content")

            yield Static("", classes="spacer")

            with Vertical(id="progress-section"):
                yield Static(
                    "Sending Progress",
                    classes="section-header",
                )

                # Batch progress bar
                yield Static("[#56b6c2][=] Batch Progress:[/]", id="batch-label")
                with Horizontal(id="batch-progress-row"):
                    yield ProgressBar(id="batch-progress", total=100, show_eta=True)
                yield Static(
                    "[#5c6370]0 / 0 batches[/]",
                    id="batch-counter",
                )

                yield Static("", classes="mini-spacer")

                # Email progress bar
                yield Static("[#56b6c2][=] Email Progress:[/]", id="email-label")
                with Horizontal(id="email-progress-row"):
                    yield ProgressBar(id="email-progress", total=100, show_eta=True)
                yield Static(
                    "[#5c6370]0 / 0 emails[/]",
                    id="email-counter",
                )

                yield Static("", classes="mini-spacer")

                # Current batch indicator
                yield Static(
                    "[#5c6370]Ready to send[/]",
                    id="progress-status",
                )

                # Stats display
                with Horizontal(id="stats-row"):
                    yield Static("[#98c379][ok] Sent: 0[/]", id="stat-sent")
                    yield Static("[#e06c75][x] Failed: 0[/]", id="stat-failed")
                    yield Static("[#e5c07b][...] Remaining: 0[/]", id="stat-remaining")

            yield Static("", classes="spacer")

            with Vertical(id="log-section"):
                yield Static(
                    "Activity Log",
                    classes="section-header",
                )
                yield Log(id="send-log", highlight=True, max_lines=100)

            yield Static("", classes="spacer")

            with Horizontal(id="send-buttons"):
                yield Button(
                    "[>] Start Sending",
                    id="btn-start",
                    classes="-primary",
                )
                yield Button(
                    "[||] Pause",
                    id="btn-pause",
                    classes="-warning",
                    disabled=True,
                )
                yield Button(
                    "[>] Resume",
                    id="btn-resume",
                    classes="-success",
                    disabled=True,
                )
                yield Button(
                    "[x] Stop",
                    id="btn-stop",
                    classes="-error",
                    disabled=True,
                )
                yield Button(
                    "[<] Back",
                    id="btn-back",
                )

            # Additional buttons for retry and export (hidden initially)
            with Horizontal(id="post-send-buttons"):
                yield Button(
                    "[~] Retry Failed",
                    id="btn-retry",
                    classes="-warning",
                    disabled=True,
                )
                yield Button(
                    "[^] Export Failed to Excel",
                    id="btn-export",
                    classes="-primary",
                    disabled=True,
                )

        yield Footer()

    def on_mount(self) -> None:
        """Initialize the send screen."""
        self._display_summary()
        self._reset_state()

    def _reset_state(self) -> None:
        """Reset sending state."""
        self.is_sending = False
        self.is_paused = False
        self.cancel_requested = False
        self.total_sent = 0
        self.total_failed = 0
        self.failed_emails = []
        self.current_batch_index = 0
        self.sending_complete = False
        self.current_email_id = None

    def _display_summary(self) -> None:
        """Display email summary."""
        email_data = getattr(self.app, "email_data", None)
        if not email_data:
            self.query_one("#summary-content", Static).update(
                "[#e06c75]No email data found. Please compose an email first.[/]"
            )
            return

        from config import get_config

        config = get_config().config

        total_recipients = len(email_data["recipients"])
        batch_size = config.batch.batch_size
        total_batches = (total_recipients + batch_size - 1) // batch_size

        summary_lines = [
            f"[#61afef]To:[/] {total_recipients} recipient(s)",
            f"[#61afef]Subject:[/] {email_data['subject']}",
            f"[#61afef]Type:[/] {email_data['email_type'].upper()}",
            f"[#61afef]Attachments:[/] {len(email_data['attachments'])} file(s)",
            "",
            f"[#61afef]Total Batches:[/] {total_batches}",
            f"[#61afef]Batch Size:[/] {batch_size} emails per batch",
            f"[#61afef]Delay:[/] {config.batch.delay_seconds}s between batches",
            f"[#61afef]Sender:[/] {format_source_email(config.aws.source_email, config.sender.sender_name)}",
        ]

        self.query_one("#summary-content", Static).update("\n".join(summary_lines))

        # Set up progress bars
        self.query_one("#batch-progress", ProgressBar).update(
            total=total_batches, progress=0
        )
        self.query_one("#email-progress", ProgressBar).update(
            total=total_recipients, progress=0
        )

        # Update counters
        self.query_one("#batch-counter", Static).update(
            f"[#5c6370]0 / {total_batches} batches[/]"
        )
        self.query_one("#email-counter", Static).update(
            f"[#5c6370]0 / {total_recipients} emails[/]"
        )
        self.query_one("#stat-remaining", Static).update(
            f"[#e5c07b][...] Remaining: {total_recipients}[/]"
        )

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button presses."""
        button_id = event.button.id

        if button_id == "btn-start":
            self._start_sending()
        elif button_id == "btn-pause":
            self._pause_sending()
        elif button_id == "btn-resume":
            self._resume_sending()
        elif button_id == "btn-stop":
            self._stop_sending()
        elif button_id == "btn-retry":
            self._retry_failed()
        elif button_id == "btn-export":
            self._export_failed_to_excel()
        elif button_id == "btn-back":
            self.action_go_back()

    def _start_sending(self) -> None:
        """Start the email sending process."""
        email_data = getattr(self.app, "email_data", None)
        if not email_data:
            self.notify("No email data found", severity="error")
            return

        if self.is_sending:
            self.notify("Already sending emails", severity="warning")
            return

        self.is_sending = True
        self.is_paused = False
        self.cancel_requested = False
        self.total_sent = 0
        self.total_failed = 0
        self.failed_emails = []
        self.current_batch_index = 0
        self.sending_complete = False

        # Update button states
        self.query_one("#btn-start", Button).disabled = True
        self.query_one("#btn-pause", Button).disabled = False
        self.query_one("#btn-resume", Button).disabled = True
        self.query_one("#btn-stop", Button).disabled = False
        self.query_one("#btn-back", Button).disabled = True
        self.query_one("#btn-retry", Button).disabled = True
        self.query_one("#btn-export", Button).disabled = True

        # Start sending in background
        self.run_worker(self._send_emails_worker(), exclusive=True)

    def _pause_sending(self) -> None:
        """Pause the sending process."""
        if not self.is_sending or self.is_paused:
            return

        self.is_paused = True
        self._log("[||] Pausing... Will pause after current batch completes.")
        self.query_one("#btn-pause", Button).disabled = True
        self._update_status("[#e5c07b][||] Pausing...[/]")

    def _resume_sending(self) -> None:
        """Resume the sending process."""
        if not self.is_paused:
            return

        self.is_paused = False
        self._log("[>] Resuming sending...")

        # Update button states
        self.query_one("#btn-pause", Button).disabled = False
        self.query_one("#btn-resume", Button).disabled = True
        self.query_one("#btn-stop", Button).disabled = False

        # Resume sending from where we left off
        self.run_worker(self._send_emails_worker(resume=True), exclusive=True)

    def _stop_sending(self) -> None:
        """Request to stop sending."""
        self.cancel_requested = True
        self.is_paused = False
        self._log("[x] Stop requested. Finishing current batch...")
        self.query_one("#btn-pause", Button).disabled = True
        self.query_one("#btn-resume", Button).disabled = True
        self.query_one("#btn-stop", Button).disabled = True
        self._update_status("[#e5c07b][x] Stopping...[/]")

    def _retry_failed(self) -> None:
        """Retry sending failed emails."""
        if not self.failed_emails:
            self.notify("No failed emails to retry", severity="warning")
            return

        if self.is_sending:
            self.notify("Already sending emails", severity="warning")
            return

        # Extract just the email addresses from failed emails
        failed_recipients = [item["email"] for item in self.failed_emails]

        self._log(f"[~] Retrying {len(failed_recipients)} failed emails...")

        # Clear failed list and reset counters
        self.failed_emails = []
        self.total_failed = 0
        self.is_sending = True
        self.is_paused = False
        self.cancel_requested = False
        self.current_batch_index = 0
        self.sending_complete = False

        # Update button states
        self.query_one("#btn-start", Button).disabled = True
        self.query_one("#btn-pause", Button).disabled = False
        self.query_one("#btn-resume", Button).disabled = True
        self.query_one("#btn-stop", Button).disabled = False
        self.query_one("#btn-back", Button).disabled = True
        self.query_one("#btn-retry", Button).disabled = True
        self.query_one("#btn-export", Button).disabled = True

        # Start retry worker
        self.run_worker(
            self._send_emails_worker(retry_recipients=failed_recipients), exclusive=True
        )

    def _export_failed_to_excel(self) -> None:
        """Export failed emails to an Excel file."""
        if not self.failed_emails:
            self.notify("No failed emails to export", severity="warning")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Failed Emails"

            # Add headers
            headers = ["Email Address", "Error Reason", "Failed At"]
            ws.append(headers)

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="E06C75", end_color="E06C75", fill_type="solid"
            )
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            # Add failed email data
            for item in self.failed_emails:
                ws.append(
                    [
                        item.get("email", "Unknown"),
                        item.get("error", "Unknown error"),
                        item.get("timestamp", datetime.datetime.now().isoformat()),
                    ]
                )

            # Auto-adjust column widths
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Save the file
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"failed_emails_{timestamp}.xlsx"
            filepath = Path("data") / filename

            # Ensure data directory exists
            filepath.parent.mkdir(parents=True, exist_ok=True)

            wb.save(filepath)

            self._log(
                f"[^] Exported {len(self.failed_emails)} failed emails to {filepath}"
            )
            self.notify(
                f"Exported to {filepath}",
                title="Export Successful",
                severity="information",
            )

        except ImportError:
            self.notify(
                "openpyxl is required for Excel export. Install with: pip install openpyxl",
                severity="error",
            )
        except Exception as e:
            self._log(f"[x] Export failed: {str(e)}")
            self.notify(f"Export failed: {str(e)}", severity="error")

    async def _send_emails_worker(
        self, resume: bool = False, retry_recipients: list[str] | None = None
    ) -> None:
        """Worker to send emails in batches."""
        from config import get_config

        email_data = self.app.email_data
        config = get_config()
        config.apply_env_vars()

        # Use retry recipients if provided, otherwise use original recipients
        if retry_recipients:
            recipients = retry_recipients
        else:
            recipients = email_data["recipients"]

        subject = email_data["subject"]
        body = email_data["body"]
        email_type = email_data["email_type"]
        attachments = email_data["attachments"]

        batch_size = config.config.batch.batch_size
        delay = config.config.batch.delay_seconds
        use_bcc = config.config.batch.use_bcc

        total_batches = (len(recipients) + batch_size - 1) // batch_size
        total_recipients = len(recipients)

        email_progress = self.query_one("#email-progress", ProgressBar)

        # Update progress bars for retry or resume
        self.query_one("#batch-progress", ProgressBar).update(
            total=total_batches, progress=self.current_batch_index
        )
        self.query_one("#email-progress", ProgressBar).update(
            total=total_recipients, progress=self.current_batch_index * batch_size
        )

        if retry_recipients:
            self._log(
                f"[~] Retrying {total_recipients} failed emails in {total_batches} batches"
            )
        elif resume:
            self._log(
                f"[>] Resuming from batch {self.current_batch_index + 1} of {total_batches}"
            )
        else:
            self._log(
                f"[>] Starting to send {total_recipients} emails in {total_batches} batches"
            )

        self._log(f"[=] Batch size: {batch_size} | Delay: {delay}s")
        self._update_status("[#56b6c2][>] Sending in progress...[/]")

        try:
            import boto3
            from botocore.exceptions import ClientError

            # Set up SES client
            ses_client = boto3.client(
                "sesv2",
                region_name=config.config.aws.region,
                aws_access_key_id=config.config.aws.access_key_id,
                aws_secret_access_key=config.config.aws.secret_access_key,
            )

            # Start from current batch index (for resume)
            start_index = self.current_batch_index * batch_size

            for batch_num, i in enumerate(
                range(start_index, len(recipients), batch_size),
                self.current_batch_index + 1,
            ):
                if self.cancel_requested:
                    self._log("[x] Sending cancelled by user")
                    break

                if self.is_paused:
                    self._log(f"[||] Paused at batch {batch_num}")
                    self._update_status(
                        f"[#e5c07b][||] Paused at batch {batch_num}/{total_batches}[/]"
                    )
                    # Save current position
                    self.current_batch_index = batch_num - 1

                    # Update button states for paused state
                    self.query_one("#btn-pause", Button).disabled = True
                    self.query_one("#btn-resume", Button).disabled = False
                    self.query_one("#btn-stop", Button).disabled = False
                    self.query_one("#btn-back", Button).disabled = True
                    return  # Exit worker, will resume later

                batch = recipients[i : i + batch_size]

                # Update batch progress
                self._update_batch_progress(batch_num, total_batches)
                self._update_status(
                    f"[#56b6c2][>] Sending batch {batch_num} of {total_batches}...[/]"
                )

                try:
                    # Create message
                    msg = self._create_message(
                        subject=subject,
                        body=body,
                        email_type=email_type,
                        attachments=attachments,
                        recipients=batch,
                        config=config.config,
                    )

                    # Send email
                    source_email = config.config.aws.source_email
                    from_address = format_source_email(
                        source_email, config.config.sender.sender_name
                    )
                    to_address = config.config.sender.default_to or get_email_address(
                        source_email
                    )
                    response = ses_client.send_email(
                        FromEmailAddress=from_address,
                        Destination={
                            "BccAddresses": batch if use_bcc else [],
                            "ToAddresses": ([to_address] if use_bcc else batch),
                        },
                        Content={
                            "Raw": {
                                "Data": msg.as_bytes(),
                            }
                        },
                    )

                    self.total_sent += len(batch)
                    self._log(
                        f"[ok] Batch {batch_num}/{total_batches}: Sent {len(batch)} emails "
                        f"(ID: {response['MessageId'][:16]}...)"
                    )

                    # Save to database
                    self._save_to_database(
                        subject=subject,
                        body=body,
                        recipients=batch,
                        attachments=attachments,
                    )

                except ClientError as e:
                    self.total_failed += len(batch)
                    error_msg = e.response["Error"]["Message"]
                    self._log(
                        f"[x] Batch {batch_num}/{total_batches}: Failed - {error_msg}"
                    )
                    # Track failed emails and save to database
                    for email_addr in batch:
                        self.failed_emails.append(
                            {
                                "email": email_addr,
                                "error": error_msg,
                                "timestamp": datetime.datetime.now().isoformat(),
                            }
                        )
                    # Save failed emails to database
                    self._save_failed_to_database(
                        recipients=batch,
                        error_reason=error_msg,
                        subject=subject,
                        body=body,
                        attachments=attachments,
                    )

                except Exception as e:
                    self.total_failed += len(batch)
                    error_msg = str(e)
                    self._log(
                        f"[x] Batch {batch_num}/{total_batches}: Error - {error_msg}"
                    )
                    # Track failed emails and save to database
                    for email_addr in batch:
                        self.failed_emails.append(
                            {
                                "email": email_addr,
                                "error": error_msg,
                                "timestamp": datetime.datetime.now().isoformat(),
                            }
                        )
                    # Save failed emails to database
                    self._save_failed_to_database(
                        recipients=batch,
                        error_reason=error_msg,
                        subject=subject,
                        body=body,
                        attachments=attachments,
                    )

                # Update current batch index
                self.current_batch_index = batch_num

                # Update email progress and stats
                emails_processed = i + len(batch)
                email_progress.update(progress=emails_processed)
                self._update_stats(total_recipients - emails_processed)
                self._update_email_counter(emails_processed, total_recipients)

                # Delay between batches (except for last batch)
                if (
                    batch_num < total_batches
                    and not self.cancel_requested
                    and not self.is_paused
                ):
                    for remaining in range(int(delay), 0, -1):
                        if self.cancel_requested or self.is_paused:
                            break
                        self._update_status(
                            f"[#e5c07b][...] Next batch in {remaining}s...[/]"
                        )
                        await asyncio.sleep(1)

        except Exception as e:
            self._log(f"[x] Critical error: {str(e)}")
            self.notify(f"Error: {str(e)}", severity="error")

        # Sending complete (only if not paused)
        if not self.is_paused:
            self.is_sending = False
            self.sending_complete = True

            if self.cancel_requested:
                self._update_status(
                    f"[#e5c07b][x] Stopped![/] Sent: {self.total_sent}, Failed: {self.total_failed}"
                )
            else:
                self._update_status(
                    f"[#98c379][ok] Complete![/] Sent: {self.total_sent}, Failed: {self.total_failed}"
                )

            self._log(
                f"[=] Final Summary: {self.total_sent} sent, {self.total_failed} failed"
            )

            # Update button states for completed state
            self.query_one("#btn-start", Button).disabled = False
            self.query_one("#btn-pause", Button).disabled = True
            self.query_one("#btn-resume", Button).disabled = True
            self.query_one("#btn-stop", Button).disabled = True
            self.query_one("#btn-back", Button).disabled = False

            # Change back button to close
            back_btn = self.query_one("#btn-back", Button)
            back_btn.label = "[x] Close"

            # Enable retry and export buttons if there are failed emails
            if self.failed_emails:
                self.query_one("#btn-retry", Button).disabled = False
                self.query_one("#btn-export", Button).disabled = False
                self._log(
                    f"[!] {len(self.failed_emails)} emails failed. Use 'Retry Failed' to resend or 'Export Failed' to save list."
                )

            if self.total_failed == 0 and not self.cancel_requested:
                self.notify(
                    f"[ok] Successfully sent {self.total_sent} emails!",
                    severity="information",
                )
            elif self.total_failed > 0:
                self.notify(
                    f"[!] Completed with {self.total_failed} failed emails",
                    severity="warning",
                )

    def _create_message(
        self,
        subject: str,
        body: str,
        email_type: str,
        attachments: list[str],
        recipients: list[str],
        config,
    ) -> MIMEMultipart:
        """Create the email message."""
        msg = MIMEMultipart()
        msg["Subject"] = subject
        source_email = config.aws.source_email
        msg["From"] = format_source_email(source_email, config.sender.sender_name)
        msg["To"] = config.sender.default_to or get_email_address(source_email)
        msg["Reply-To"] = config.sender.reply_to or get_email_address(source_email)

        # Add body
        body_part = MIMEMultipart("alternative")
        mime_type = "html" if email_type == "html" else "plain"
        body_part.attach(MIMEText(body, mime_type))
        msg.attach(body_part)

        # Add attachments
        for attachment_path in attachments:
            try:
                path = Path(attachment_path)
                with open(path, "rb") as f:
                    part = MIMEApplication(f.read(), Name=path.name)
                    part.add_header(
                        "Content-Disposition",
                        "attachment",
                        filename=path.name,
                    )
                    msg.attach(part)
            except Exception as e:
                self._log(f"[!] Could not attach {attachment_path}: {e}")

        return msg

    def _save_to_database(
        self,
        subject: str,
        body: str,
        recipients: list[str],
        attachments: list[str],
    ) -> None:
        """Save sent email record to database."""
        try:
            import sys
            from pathlib import Path

            # Add sending directory to path
            sys.path.insert(0, str(Path(__file__).parent.parent.parent / "sending"))
            from db import Database
            from emails import Email

            from config import get_config

            config = get_config().config

            db = Database()

            # Create or reuse email record
            if self.current_email_id:
                email_id = self.current_email_id
            else:
                # Create email record
                email = Email(
                    body=body,
                    subject=subject,
                    sender=config.sender.sender_name,
                    recipient=config.aws.source_email,
                    files=attachments,
                )
                email_id = email.id
                self.current_email_id = email_id

                # Add email to database
                try:
                    db.add_email(email)
                except Exception:
                    # Email might already exist
                    pass

            # Add sent records
            for recipient in recipients:
                db.add_sent(email_id, recipient, "bcc")

            db.close()

        except Exception as e:
            self._log(f"[!] Could not save to database: {e}")

    def _save_failed_to_database(
        self,
        recipients: list[str],
        error_reason: str,
        subject: str,
        body: str,
        attachments: list[str],
    ) -> None:
        """Save failed email records to database."""
        try:
            import sys
            from pathlib import Path

            # Add sending directory to path
            sys.path.insert(0, str(Path(__file__).parent.parent.parent / "sending"))
            from db import Database
            from emails import Email

            from config import get_config

            config = get_config().config

            db = Database()

            # Create or reuse email record
            if self.current_email_id:
                email_id = self.current_email_id
            else:
                # Create email record
                email = Email(
                    body=body,
                    subject=subject,
                    sender=config.sender.sender_name,
                    recipient=config.aws.source_email,
                    files=attachments,
                )
                email_id = email.id
                self.current_email_id = email_id

                # Add email to database
                try:
                    db.add_email(email)
                except Exception:
                    # Email might already exist
                    pass

            # Add failed records
            for recipient in recipients:
                db.add_failed_email(email_id, recipient, error_reason)

            db.close()

        except Exception as e:
            self._log(f"[!] Could not save failed emails to database: {e}")

    def _log(self, message: str) -> None:
        """Add a message to the activity log."""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        log_widget = self.query_one("#send-log", Log)
        log_widget.write_line(f"[{timestamp}] {message}")

    def _update_status(self, status: str) -> None:
        """Update the progress status."""
        self.query_one("#progress-status", Static).update(status)

    def _update_batch_progress(self, current_batch: int, total_batches: int) -> None:
        """Update batch progress bar and counter."""
        batch_progress = self.query_one("#batch-progress", ProgressBar)
        batch_progress.update(progress=current_batch)

        percentage = (current_batch / total_batches) * 100 if total_batches > 0 else 0
        self.query_one("#batch-counter", Static).update(
            f"[#56b6c2]{current_batch} / {total_batches} batches[/] ({percentage:.1f}%)"
        )

    def _update_email_counter(self, sent: int, total: int) -> None:
        """Update email counter display."""
        percentage = (sent / total) * 100 if total > 0 else 0
        self.query_one("#email-counter", Static).update(
            f"[#56b6c2]{sent} / {total} emails[/] ({percentage:.1f}%)"
        )

    def _update_stats(self, remaining: int) -> None:
        """Update statistics display."""
        self.query_one("#stat-sent", Static).update(
            f"[#98c379][ok] Sent: {self.total_sent}[/]"
        )
        self.query_one("#stat-failed", Static).update(
            f"[#e06c75][x] Failed: {self.total_failed}[/]"
        )
        self.query_one("#stat-remaining", Static).update(
            f"[#e5c07b][...] Remaining: {remaining}[/]"
        )

    def action_cancel(self) -> None:
        """Handle escape key."""
        if self.is_sending and not self.is_paused:
            self._pause_sending()
        elif self.is_paused:
            self._stop_sending()
        else:
            self.action_go_back()

    def action_go_back(self) -> None:
        """Go back or close based on sending state."""
        if self.is_sending and not self.is_paused:
            self.notify(
                "Cannot go back while sending. Pause or stop sending first.",
                severity="warning",
            )
            return

        if self.is_paused:
            self.notify(
                "Sending is paused. Stop sending first to go back.",
                severity="warning",
            )
            return

        # If sending is complete, go back to home instead of compose
        if self.sending_complete:
            # Pop all screens and go home
            while len(self.app.screen_stack) > 1:
                self.app.pop_screen()
            if self.app.screen.name != "home":
                # Use push_screen instead of switch_screen to avoid callback stack error
                self.app.push_screen("home")
        else:
            self.app.pop_screen()
