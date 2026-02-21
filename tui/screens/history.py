"""
History Screen for SES Email TUI application.
Displays sent email records from the database with a master-detail view.
"""

import datetime
from pathlib import Path
from typing import Optional

from textual.app import ComposeResult
from textual.containers import Container, Horizontal, ScrollableContainer, Vertical
from textual.screen import Screen
from textual.widgets import (
    Button,
    DataTable,
    Footer,
    Header,
    Input,
    Static,
)


class HistoryScreen(Screen):
    """Screen for viewing email history with master-detail layout."""

    BINDINGS = [
        ("escape", "go_back", "Back"),
        ("r", "refresh", "Refresh"),
        ("/", "focus_search", "Search"),
    ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.selected_email_id: Optional[str] = None
        self.email_stats: dict = {}  # Cache for email statistics

    def compose(self) -> ComposeResult:
        yield Header()
        with Container(id="history-container"):
            yield Static(
                "[bold]#[/] Email History",
                classes="screen-title",
            )

            # Search bar
            with Horizontal(id="search-row"):
                yield Input(
                    placeholder="[?] Search by subject or sender...",
                    id="search-input",
                )
                yield Button("[?] Search", id="btn-search", classes="-primary")
                yield Button("[~] Refresh", id="btn-refresh")

            # Master section - Email list
            with Vertical(id="emails-section"):
                yield Static(
                    "[bold #56b6c2]Sent Emails[/]",
                    classes="section-header",
                )
                yield DataTable(id="emails-table", cursor_type="row")
                yield Static("", id="emails-count")

            # Detail section - Shows when email is selected
            with Vertical(id="detail-section"):
                yield Static(
                    "[bold #56b6c2]Email Details[/]",
                    classes="section-header",
                )
                with ScrollableContainer(id="detail-content-wrapper"):
                    yield Static(
                        "[#5c6370]Select an email above to view details[/]",
                        id="detail-content",
                    )

                # Action buttons for selected email
                with Horizontal(id="detail-actions"):
                    yield Button(
                        "[+] Resend to New",
                        id="btn-resend-new",
                        classes="-primary",
                        disabled=True,
                    )
                    yield Button(
                        "[~] Retry Failed",
                        id="btn-retry-failed",
                        classes="-warning",
                        disabled=True,
                    )
                    yield Button(
                        "[^] Export Failed",
                        id="btn-export-failed",
                        classes="-primary",
                        disabled=True,
                    )
                    yield Button(
                        "[=] Export All",
                        id="btn-export-all",
                        classes="-success",
                        disabled=True,
                    )

            # Bottom buttons
            with Horizontal(id="history-buttons"):
                yield Button("[%] Statistics", id="btn-stats")
                yield Button("[<] Back", id="btn-back")

        yield Footer()

    def on_mount(self) -> None:
        """Initialize the history screen."""
        self._setup_table()
        self._load_data()

    def _setup_table(self) -> None:
        """Set up the data table with columns."""
        emails_table = self.query_one("#emails-table", DataTable)
        emails_table.add_column("Subject", key="subject", width=35)
        emails_table.add_column("Sender", key="sender", width=20)
        emails_table.add_column("Last Sent", key="last_sent", width=18)
        emails_table.add_column("Sent", key="sent", width=8)
        emails_table.add_column("Failed", key="failed", width=8)
        emails_table.add_column("Status", key="status", width=10)
        emails_table.zebra_stripes = True

    def _load_data(self, search_term: str = "") -> None:
        """Load data from the database using grouped emails."""
        try:
            import sys

            sys.path.insert(0, str(Path(__file__).parent.parent.parent / "sending"))
            from db import Database

            db = Database()

            # Get grouped emails (consolidates duplicates with same subject/sender/body)
            grouped_emails = db.get_grouped_emails_summary()
            self.email_stats = {}

            for group in grouped_emails:
                group_id = group["id"]  # Primary ID for the group
                email_ids = group["email_ids"]  # All email IDs in this group

                # Get all sent and failed records for all templates in this group
                sent_records = db.get_sent_emails_by_email_ids(email_ids)
                failed_records = db.get_failed_emails_by_email_ids(email_ids)

                # Create a pseudo-email tuple for compatibility with existing code
                # Format: (id, subject, body, sender, files)
                email_tuple = (
                    group_id,
                    group["subject"],
                    group["body"],
                    group["sender"],
                    group["files"],
                )

                self.email_stats[group_id] = {
                    "email": email_tuple,
                    "email_ids": email_ids,  # All template IDs in this group
                    "sent_count": group["sent_count"],
                    "failed_count": group["failed_count"],
                    "sent_records": sent_records,
                    "failed_records": failed_records,
                    "last_sent": group["last_sent"],
                    "template_count": group["template_count"],
                }

            self._populate_table(grouped_emails, search_term)
            db.close()

        except Exception as e:
            self.notify(f"Error loading data: {e}", severity="error")

    def _populate_table(self, grouped_emails: list, search_term: str = "") -> None:
        """Populate the emails table with grouped/consolidated data."""
        emails_table = self.query_one("#emails-table", DataTable)
        emails_table.clear()

        # Filter emails based on search
        filtered_emails = grouped_emails
        if search_term:
            search_lower = search_term.lower()
            filtered_emails = [
                e
                for e in grouped_emails
                if search_lower in str(e["subject"]).lower()
                or search_lower in str(e["sender"]).lower()
            ]

        # Already sorted by last_sent from the database method
        for group in filtered_emails:
            group_id = group["id"]
            stats = self.email_stats.get(group_id, {})

            subject = (
                group["subject"][:33] + "..."
                if len(group["subject"]) > 33
                else group["subject"]
            )
            sender = (
                group["sender"][:18] + "..."
                if len(group["sender"]) > 18
                else group["sender"]
            )

            sent_count = group["sent_count"]
            failed_count = group["failed_count"]

            # Format last sent date
            last_sent = group["last_sent"]
            if last_sent:
                last_sent_str = str(last_sent)[:16]
            else:
                last_sent_str = "-"

            # Determine status
            if sent_count == 0 and failed_count == 0:
                status = "[#5c6370]Draft[/]"
            elif failed_count == 0:
                status = "[#98c379]Success[/]"
            elif sent_count == 0:
                status = "[#e06c75]Failed[/]"
            else:
                status = "[#e5c07b]Partial[/]"

            emails_table.add_row(
                subject,
                sender,
                last_sent_str,
                str(sent_count),
                str(failed_count) if failed_count > 0 else "-",
                status,
                key=group_id,
            )

        count = len(filtered_emails)
        total = len(grouped_emails)
        if search_term:
            self.query_one("#emails-count", Static).update(
                f"[#5c6370]Showing {count} of {total} email campaigns[/]"
            )
        else:
            self.query_one("#emails-count", Static).update(
                f"[#5c6370]{total} email campaign(s) in history[/]"
            )

    def _show_email_details(self, email_id: str) -> None:
        """Show details for the selected email."""
        try:
            stats = self.email_stats.get(email_id)
            if not stats:
                self.query_one("#detail-content", Static).update(
                    "[#e06c75]Email data not found. Try refreshing.[/]"
                )
                return

            email = stats["email"]
            sent_count = stats["sent_count"]
            failed_count = stats["failed_count"]
            sent_records = stats["sent_records"]
            failed_records = stats["failed_records"]
            last_sent = stats["last_sent"]

            # Build detail content
            lines = []

            # Header with stats
            lines.append("[bold #c678dd]═══ Email Summary ═══[/]")
            lines.append("")

            # Status bar
            total = sent_count + failed_count
            if total > 0:
                success_rate = (sent_count / total) * 100
                if success_rate == 100:
                    status_color = "#98c379"
                    status_text = "All Successful"
                elif success_rate == 0:
                    status_color = "#e06c75"
                    status_text = "All Failed"
                else:
                    status_color = "#e5c07b"
                    status_text = f"{success_rate:.1f}% Success"
                lines.append(f"[bold {status_color}]Status: {status_text}[/]")
            lines.append("")

            # Stats grid
            lines.append(
                f"[#98c379]✓ Sent:[/] {sent_count}    [#e06c75]✗ Failed:[/] {failed_count}    [#61afef]Total:[/] {total}"
            )
            lines.append("")

            # Template details
            lines.append("[bold #56b6c2]─── Template Details ───[/]")
            lines.append("")
            lines.append(f"[#61afef]Subject:[/] {email[1]}")
            lines.append(f"[#61afef]Sender:[/] {email[3]}")

            # Show template count if there are multiple
            template_count = stats.get("template_count", 1)
            if template_count > 1:
                lines.append(
                    f"[#e5c07b]Templates:[/] {template_count} batches consolidated"
                )

            # Attachments
            attachments = email[4] if len(email) > 4 and email[4] else None
            if attachments:
                attachment_list = [
                    a.strip() for a in attachments.split(",") if a.strip()
                ]
                lines.append(f"[#61afef]Attachments:[/] {len(attachment_list)} file(s)")
                for att in attachment_list[:3]:
                    lines.append(f"  [#5c6370]• {Path(att).name}[/]")
                if len(attachment_list) > 3:
                    lines.append(
                        f"  [#5c6370]... and {len(attachment_list) - 3} more[/]"
                    )
            else:
                lines.append(f"[#61afef]Attachments:[/] None")

            if last_sent:
                lines.append(f"[#61afef]Last Sent:[/] {str(last_sent)[:19]}")
            lines.append("")

            # Body preview
            lines.append("[bold #56b6c2]─── Body Preview ───[/]")
            lines.append("")
            body_preview = email[2][:400] + "..." if len(email[2]) > 400 else email[2]
            # Escape any markup in body
            body_preview = body_preview.replace("[", "\\[").replace("]", "\\]")
            lines.append(f"[#abb2bf]{body_preview}[/]")
            lines.append("")

            # Failed recipients section
            if failed_records:
                lines.append("[bold #e06c75]─── Failed Recipients ───[/]")
                lines.append("")
                # Group by error
                errors = {}
                for record in failed_records:
                    error = record[3] if len(record) > 3 else "Unknown error"
                    if error not in errors:
                        errors[error] = []
                    errors[error].append(record[2])  # recipient

                for error, recipients in errors.items():
                    error_short = error[:60] + "..." if len(error) > 60 else error
                    lines.append(f"[#e06c75]Error:[/] {error_short}")
                    for recipient in recipients[:5]:
                        lines.append(f"  [#e06c75]✗[/] {recipient}")
                    if len(recipients) > 5:
                        lines.append(
                            f"  [#5c6370]... and {len(recipients) - 5} more[/]"
                        )
                    lines.append("")

            # Recent successful recipients
            if sent_records:
                lines.append("[bold #98c379]─── Recent Recipients ───[/]")
                lines.append("")
                # Show most recent 10
                recent = sorted(
                    sent_records, key=lambda r: r[4] if len(r) > 4 else "", reverse=True
                )[:10]
                for record in recent:
                    recipient = record[2]
                    sent_at = str(record[4])[:16] if len(record) > 4 else ""
                    lines.append(f"  [#98c379]✓[/] {recipient} [#5c6370]({sent_at})[/]")
                if len(sent_records) > 10:
                    lines.append(f"  [#5c6370]... and {len(sent_records) - 10} more[/]")

            self.query_one("#detail-content", Static).update("\n".join(lines))
            self.selected_email_id = email_id

            # Enable action buttons
            self.query_one("#btn-export-all", Button).disabled = False
            self.query_one("#btn-resend-new", Button).disabled = False
            if failed_count > 0:
                self.query_one("#btn-retry-failed", Button).disabled = False
                self.query_one("#btn-export-failed", Button).disabled = False
            else:
                self.query_one("#btn-retry-failed", Button).disabled = True
                self.query_one("#btn-export-failed", Button).disabled = True

        except Exception as e:
            self.query_one("#detail-content", Static).update(
                f"[#e06c75]Error loading details: {e}[/]"
            )

    def _show_statistics(self) -> None:
        """Show overall statistics."""
        try:
            total_emails = len(self.email_stats)
            total_sent = sum(s["sent_count"] for s in self.email_stats.values())
            total_failed = sum(s["failed_count"] for s in self.email_stats.values())

            # Count unique recipients
            all_recipients = set()
            for stats in self.email_stats.values():
                for record in stats["sent_records"]:
                    all_recipients.add(record[2])

            # Today's stats
            today = datetime.date.today()
            today_sent = 0
            week_sent = 0
            for stats in self.email_stats.values():
                for record in stats["sent_records"]:
                    if len(record) > 4 and record[4]:
                        try:
                            sent_date = datetime.datetime.fromisoformat(
                                str(record[4]).replace(" ", "T")
                            ).date()
                            if sent_date == today:
                                today_sent += 1
                            if (today - sent_date).days <= 7:
                                week_sent += 1
                        except (ValueError, TypeError):
                            pass

            lines = [
                "[bold #c678dd]═══ Email Statistics ═══[/]",
                "",
                "[bold #56b6c2]Overall[/]",
                "",
                f"  [#61afef]Email Templates:[/] {total_emails}",
                f"  [#98c379]Total Sent:[/] {total_sent}",
                f"  [#e06c75]Total Failed:[/] {total_failed}",
                f"  [#61afef]Unique Recipients:[/] {len(all_recipients)}",
                "",
                "[bold #56b6c2]Recent Activity[/]",
                "",
                f"  [#e5c07b]Today:[/] {today_sent} emails sent",
                f"  [#e5c07b]This Week:[/] {week_sent} emails sent",
                "",
                "[bold #56b6c2]Success Rate[/]",
                "",
            ]

            total = total_sent + total_failed
            if total > 0:
                success_rate = (total_sent / total) * 100
                lines.append(f"  [#98c379]Success Rate:[/] {success_rate:.1f}%")

                # Visual bar
                bar_width = 30
                filled = int(bar_width * success_rate / 100)
                bar = (
                    "[#98c379]"
                    + "█" * filled
                    + "[/][#e06c75]"
                    + "█" * (bar_width - filled)
                    + "[/]"
                )
                lines.append(f"  {bar}")
            else:
                lines.append("  [#5c6370]No emails sent yet[/]")

            self.query_one("#detail-content", Static).update("\n".join(lines))

            # Disable action buttons for stats view
            self.query_one("#btn-retry-failed", Button).disabled = True
            self.query_one("#btn-export-failed", Button).disabled = True
            self.query_one("#btn-export-all", Button).disabled = True
            self.query_one("#btn-resend-new", Button).disabled = True
            self.selected_email_id = None

        except Exception as e:
            self.query_one("#detail-content", Static).update(
                f"[#e06c75]Error loading statistics: {e}[/]"
            )

    def _export_failed_emails(self) -> None:
        """Export failed emails for the selected email to Excel."""
        if not self.selected_email_id:
            self.notify("No email selected", severity="warning")
            return

        stats = self.email_stats.get(self.selected_email_id)
        if not stats or not stats["failed_records"]:
            self.notify("No failed emails to export", severity="warning")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Failed Emails"

            # Add headers
            headers = ["Recipient", "Error Reason", "Failed At", "Retried"]
            ws.append(headers)

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="E06C75", end_color="E06C75", fill_type="solid"
            )
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

            # Add data
            for record in stats["failed_records"]:
                ws.append(
                    [
                        record[2],  # recipient
                        record[3],  # error_reason
                        str(record[4]) if len(record) > 4 else "",  # failed_at
                        "Yes"
                        if len(record) > 5 and record[5] == 1
                        else "No",  # retried
                    ]
                )

            # Auto-adjust column widths
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            # Save file
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"failed_emails_{timestamp}.xlsx"
            filepath = Path("data") / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            wb.save(filepath)

            self.notify(
                f"Exported to {filepath}",
                title="Export Successful",
                severity="information",
            )

        except ImportError:
            self.notify(
                "openpyxl required. Install with: pip install openpyxl",
                severity="error",
            )
        except Exception as e:
            self.notify(f"Export failed: {e}", severity="error")

    def _export_all_recipients(self) -> None:
        """Export all recipients for the selected email to Excel."""
        if not self.selected_email_id:
            self.notify("No email selected", severity="warning")
            return

        stats = self.email_stats.get(self.selected_email_id)
        if not stats:
            self.notify("No data to export", severity="warning")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()

            # Sheet 1: Successful
            ws_sent = wb.active
            ws_sent.title = "Sent Successfully"

            headers_sent = ["Recipient", "Type", "Sent At"]
            ws_sent.append(headers_sent)

            header_font = Font(bold=True, color="FFFFFF")
            success_fill = PatternFill(
                start_color="98C379", end_color="98C379", fill_type="solid"
            )
            for col in range(1, len(headers_sent) + 1):
                cell = ws_sent.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = success_fill

            for record in stats["sent_records"]:
                ws_sent.append(
                    [
                        record[2],  # recipient
                        record[3],  # type
                        str(record[4]) if len(record) > 4 else "",  # sent_at
                    ]
                )

            # Sheet 2: Failed
            ws_failed = wb.create_sheet("Failed")

            headers_failed = ["Recipient", "Error Reason", "Failed At"]
            ws_failed.append(headers_failed)

            failed_fill = PatternFill(
                start_color="E06C75", end_color="E06C75", fill_type="solid"
            )
            for col in range(1, len(headers_failed) + 1):
                cell = ws_failed.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = failed_fill

            for record in stats["failed_records"]:
                ws_failed.append(
                    [
                        record[2],  # recipient
                        record[3],  # error_reason
                        str(record[4]) if len(record) > 4 else "",  # failed_at
                    ]
                )

            # Auto-adjust columns for both sheets
            for ws in [ws_sent, ws_failed]:
                for column_cells in ws.columns:
                    max_length = 0
                    column = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (TypeError, AttributeError):
                            pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)

            # Save file
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            email = stats["email"]
            subject_clean = "".join(
                c for c in email[1][:20] if c.isalnum() or c in " -_"
            )
            filename = f"recipients_{subject_clean}_{timestamp}.xlsx"
            filepath = Path("data") / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            wb.save(filepath)

            total = len(stats["sent_records"]) + len(stats["failed_records"])
            self.notify(
                f"Exported {total} recipients to {filepath}",
                title="Export Successful",
                severity="information",
            )

        except ImportError:
            self.notify(
                "openpyxl required. Install with: pip install openpyxl",
                severity="error",
            )
        except Exception as e:
            self.notify(f"Export failed: {e}", severity="error")

    def _retry_failed_emails(self) -> None:
        """Set up retry for failed emails - navigates to send screen with failed recipients."""
        if not self.selected_email_id:
            self.notify("No email selected", severity="warning")
            return

        stats = self.email_stats.get(self.selected_email_id)
        if not stats or not stats["failed_records"]:
            self.notify("No failed emails to retry", severity="warning")
            return

        # Get failed recipients
        failed_recipients = [record[2] for record in stats["failed_records"]]
        email = stats["email"]

        # Set up email data for retry
        self.app.email_data = {
            "recipients": failed_recipients,
            "subject": email[1],
            "body": email[2],
            "email_type": "html",  # Default to HTML
            "attachments": [a.strip() for a in email[4].split(",") if a.strip()]
            if email[4]
            else [],
        }

        # Mark failed emails as being retried
        try:
            import sys

            sys.path.insert(0, str(Path(__file__).parent.parent.parent / "sending"))
            from db import Database

            db = Database()
            for record in stats["failed_records"]:
                db.mark_failed_email_retried(record[0])
            db.close()
        except Exception:
            pass  # Non-critical

        self.notify(
            f"Retrying {len(failed_recipients)} failed recipients...",
            severity="information",
        )
        self.app.push_screen("send")

    def _resend_to_new_recipients(self) -> None:
        """Set up compose screen to send to new recipients for this email template."""
        if not self.selected_email_id:
            self.notify("No email selected", severity="warning")
            return

        stats = self.email_stats.get(self.selected_email_id)
        if not stats:
            self.notify("Email data not found", severity="error")
            return

        email = stats["email"]
        # Get all email IDs in this group for comparison purposes
        email_ids = stats.get("email_ids", [self.selected_email_id])

        # Store template data for the compose screen
        # The user will load their new recipient list in compose
        # and can use Compare to filter out already-sent recipients
        self.app.email_data = {
            "template_email_id": self.selected_email_id,
            "template_email_ids": email_ids,  # All IDs in this group for comparison
            "subject": email[1],
            "body": email[2],
            "email_type": "html",
            "attachments": [a.strip() for a in email[4].split(",") if a.strip()]
            if email[4]
            else [],
            "recipients": [],  # User will load new list
        }

        self.notify(
            "Opening compose with template. Load your new recipient list and use 'Compare with Previous' to filter.",
            severity="information",
            timeout=5,
        )
        self.app.push_screen("compose")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button presses."""
        button_id = event.button.id

        if button_id == "btn-search":
            self._do_search()
        elif button_id == "btn-refresh":
            self.action_refresh()
        elif button_id == "btn-stats":
            self._show_statistics()
        elif button_id == "btn-back":
            self.action_go_back()
        elif button_id == "btn-retry-failed":
            self._retry_failed_emails()
        elif button_id == "btn-export-failed":
            self._export_failed_emails()
        elif button_id == "btn-export-all":
            self._export_all_recipients()
        elif button_id == "btn-resend-new":
            self._resend_to_new_recipients()

    def on_input_submitted(self, event: Input.Submitted) -> None:
        """Handle Enter key in search input."""
        if event.input.id == "search-input":
            self._do_search()

    def on_data_table_row_selected(self, event: DataTable.RowSelected) -> None:
        """Handle row selection in data table."""
        row_key = event.row_key
        key_value = row_key.value if hasattr(row_key, "value") else str(row_key)
        self._show_email_details(key_value)

    def on_data_table_row_highlighted(self, event: DataTable.RowHighlighted) -> None:
        """Handle row highlight (cursor movement) in data table."""
        row_key = event.row_key
        if row_key:
            key_value = row_key.value if hasattr(row_key, "value") else str(row_key)
            self._show_email_details(key_value)

    def _do_search(self) -> None:
        """Perform search."""
        search_term = self.query_one("#search-input", Input).value
        self._load_data(search_term)
        if search_term:
            self.notify(f"Searching for: {search_term}", severity="information")

    def action_refresh(self) -> None:
        """Refresh the data."""
        self._load_data()
        self.query_one("#detail-content", Static).update(
            "[#5c6370]Select an email above to view details[/]"
        )
        self.query_one("#btn-retry-failed", Button).disabled = True
        self.query_one("#btn-export-failed", Button).disabled = True
        self.query_one("#btn-export-all", Button).disabled = True
        self.query_one("#btn-resend-new", Button).disabled = True
        self.selected_email_id = None
        self.notify("Data refreshed", severity="information")

    def action_focus_search(self) -> None:
        """Focus the search input."""
        self.query_one("#search-input", Input).focus()

    def action_go_back(self) -> None:
        """Go back to home screen."""
        self.app.pop_screen()
